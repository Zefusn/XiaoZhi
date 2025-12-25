from flask import Flask, request, jsonify
from flask_cors import CORS
import sqlite3
import os
from werkzeug.utils import secure_filename
import json
import sys
import csv
from io import StringIO
import time
from pathlib import Path

app = Flask(__name__)
CORS(app)

# 配置上传文件夹
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# 全局数据存储
data_store = {
    'page1_data': None,
    'page2_data': None,
    'conn': None
}


def allowed_file(filename):
    if not '.' in filename:
        return False  # 没有扩展名的文件不被允许
    ext = filename.rsplit('.', 1)[1].lower()
    return ext in {'xlsx', 'xls', 'xlsm', 'xltx', 'xltm'}


def cleanup_old_files():
    """清理超过1小时的旧文件"""
    uploads_path = Path(UPLOAD_FOLDER)
    current_time = time.time()
    for file_path in uploads_path.iterdir():
        if file_path.is_file():
            # 检查文件修改时间
            if current_time - file_path.stat().st_mtime > 3600:  # 1小时
                try:
                    file_path.unlink()
                except OSError:
                    pass  # 如果文件被占用，跳过删除


def read_excel_with_python_libs(filepath):
    """使用 openpyxl 或 xlrd 读取 Excel 文件，支持 .xlsx 和 .xls 格式"""
    import os
    file_ext = os.path.splitext(filepath)[1].lower()
    
    # 检查扩展名是否为空或无效
    if not file_ext:
        raise ValueError("文件没有扩展名，请确保上传的文件有正确的 Excel 扩展名 (.xlsx, .xls, .xlsm, .xltx, .xltm)")
    
    if file_ext == '.xlsx' or file_ext == '.xlsm' or file_ext == '.xltx' or file_ext == '.xltm':
        # 使用 openpyxl 处理 .xlsx, .xlsm, .xltx, .xltm 格式
        from openpyxl import load_workbook
        
        wb = None
        try:
            # 使用只读模式确保文件被正确处理
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active
            
            # 读取所有数据
            data = []
            headers = []
            
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    # 第一行作为列名
                    headers = [str(cell) if cell is not None else f'Column_{i}' for i, cell in enumerate(row)]
                else:
                    # 数据行
                    data.append({headers[i]: (cell if cell is not None else '') for i, cell in enumerate(row)})
            
            return data, headers
        except Exception as e:
            raise ValueError(f"无法使用 openpyxl 读取文件: {str(e)}")
        finally:
            # 确保工作簿被关闭
            if wb is not None:
                try:
                    wb.close()
                except:
                    pass
    elif file_ext == '.xls':
        # 使用 xlrd 处理 .xls 格式
        import xlrd
        
        wb = None
        try:
            # 打开工作簿
            wb = xlrd.open_workbook(filepath)
            ws = wb.sheet_by_index(0)  # 使用第一个工作表
            
            # 读取所有数据
            data = []
            headers = []
            
            for row_idx in range(ws.nrows):
                row_values = ws.row_values(row_idx)
                if row_idx == 0:
                    # 第一行作为列名
                    headers = [str(cell) if cell != '' else f'Column_{i}' for i, cell in enumerate(row_values)]
                else:
                    # 数据行
                    data.append({headers[i]: (cell if cell != '' else '') for i, cell in enumerate(row_values)})
            
            return data, headers
        except Exception as e:
            raise ValueError(f"无法使用 xlrd 读取文件: {str(e)}")
        finally:
            # xlrd 通常不需要显式关闭文件，但这里保持一致
            pass
    else:
        # 尝试检测文件的实际格式，但首先检查 magic 库是否可用
        try:
            import magic
            try:
                mime = magic.from_file(filepath, mime=True)
                if mime in ['application/vnd.ms-excel', 'application/xls']:
                    # 如果是 .xls 格式但扩展名不对，尝试用 xlrd 读取
                    import xlrd
                    wb = xlrd.open_workbook(filepath)
                    ws = wb.sheet_by_index(0)
                    
                    # 读取所有数据
                    data = []
                    headers = []
                    
                    for row_idx in range(ws.nrows):
                        row_values = ws.row_values(row_idx)
                        if row_idx == 0:
                            # 第一行作为列名
                            headers = [str(cell) if cell != '' else f'Column_{i}' for i, cell in enumerate(row_values)]
                        else:
                            # 数据行
                            data.append({headers[i]: (cell if cell != '' else '') for i, cell in enumerate(row_values)})
                    
                    return data, headers
                elif mime in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                              'application/vnd.ms-excel.sheet.macroEnabled.12', 
                              'application/vnd.openxmlformats-officedocument.spreadsheetml.template',
                              'application/vnd.ms-excel.template.macroEnabled.12']:
                    # 如果是 .xlsx 格式但扩展名不对，尝试用 openpyxl 读取
                    from openpyxl import load_workbook
                    wb = None
                    try:
                        wb = load_workbook(filepath, read_only=True)
                        ws = wb.active
                        
                        # 读取所有数据
                        data = []
                        headers = []
                        
                        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                            if row_idx == 0:
                                # 第一行作为列名
                                headers = [str(cell) if cell is not None else f'Column_{i}' for i, cell in enumerate(row)]
                            else:
                                # 数据行
                                data.append({headers[i]: (cell if cell is not None else '') for i, cell in enumerate(row)})
                        
                        return data, headers
                    finally:
                        if wb is not None:
                            try:
                                wb.close()
                            except:
                                pass
                else:
                    raise ValueError(f"不支持的文件格式: {file_ext}，MIME类型: {mime}")
            except ImportError:
                # 如果没有安装 python-magic，使用扩展名检测
                raise ValueError(f"不支持的文件格式: {file_ext}。请确保文件是 Excel 格式 (.xlsx, .xls, .xlsm, .xltx, .xltm)")
        except ImportError:
            # 如果 magic 模块不可用，直接使用扩展名检测
            raise ValueError(f"不支持的文件格式: {file_ext}。请确保文件是 Excel 格式 (.xlsx, .xls, .xlsm, .xltx, .xltm)")


def read_excel_with_python_libs_for_filter(filepath):
    """使用 openpyxl 或 xlrd 读取用于过滤的 Excel 文件，返回 userContent 列的值集合，支持 .xlsx 和 .xls 格式"""
    import os
    file_ext = os.path.splitext(filepath)[1].lower()
    
    # 检查扩展名是否为空或无效
    if not file_ext:
        raise ValueError("文件没有扩展名，请确保上传的文件有正确的 Excel 扩展名 (.xlsx, .xls, .xlsm, .xltx, .xltm)")
    
    if file_ext == '.xlsx' or file_ext == '.xlsm' or file_ext == '.xltx' or file_ext == '.xltm':
        # 使用 openpyxl 处理 .xlsx, .xlsm, .xltx, .xltm 格式
        from openpyxl import load_workbook
        
        wb = None
        try:
            # 使用只读模式确保文件被正确处理
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active
            
            # 获取所有 userContent 值
            user_content_values = set()
            headers = []
            
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    # 第一行作为列名
                    headers = [str(cell) if cell is not None else f'Column_{i}' for i, cell in enumerate(row)]
                else:
                    # 数据行
                    row_dict = {headers[i]: (cell if cell is not None else '') for i, cell in enumerate(row)}
                    if 'userContent' in row_dict and row_dict['userContent']:
                        user_content_values.add(str(row_dict['userContent']))
            
            return user_content_values
        except Exception as e:
            raise ValueError(f"无法使用 openpyxl 读取文件: {str(e)}")
        finally:
            # 确保工作簿被关闭
            if wb is not None:
                try:
                    wb.close()
                except:
                    pass
    elif file_ext == '.xls':
        # 使用 xlrd 处理 .xls 格式
        import xlrd
        
        wb = None
        try:
            # 打开工作簿
            wb = xlrd.open_workbook(filepath)
            ws = wb.sheet_by_index(0)  # 使用第一个工作表
            
            # 获取所有 userContent 值
            user_content_values = set()
            headers = []
            
            for row_idx in range(ws.nrows):
                row_values = ws.row_values(row_idx)
                if row_idx == 0:
                    # 第一行作为列名
                    headers = [str(cell) if cell != '' else f'Column_{i}' for i, cell in enumerate(row_values)]
                else:
                    # 数据行
                    row_dict = {headers[i]: (cell if cell != '' else '') for i, cell in enumerate(row_values)}
                    if 'userContent' in row_dict and row_dict['userContent']:
                        user_content_values.add(str(row_dict['userContent']))
            
            return user_content_values
        except Exception as e:
            raise ValueError(f"无法使用 xlrd 读取文件: {str(e)}")
        finally:
            # xlrd 通常不需要显式关闭文件，但这里保持一致
            pass
    else:
        # 尝试检测文件的实际格式，但首先检查 magic 库是否可用
        try:
            import magic
            try:
                mime = magic.from_file(filepath, mime=True)
                if mime in ['application/vnd.ms-excel', 'application/xls']:
                    # 如果是 .xls 格式但扩展名不对，尝试用 xlrd 读取
                    import xlrd
                    wb = xlrd.open_workbook(filepath)
                    ws = wb.sheet_by_index(0)
                    
                    # 获取所有 userContent 值
                    user_content_values = set()
                    headers = []
                    
                    for row_idx in range(ws.nrows):
                        row_values = ws.row_values(row_idx)
                        if row_idx == 0:
                            # 第一行作为列名
                            headers = [str(cell) if cell != '' else f'Column_{i}' for i, cell in enumerate(row_values)]
                        else:
                            # 数据行
                            row_dict = {headers[i]: (cell if cell != '' else '') for i, cell in enumerate(row_values)}
                            if 'userContent' in row_dict and row_dict['userContent']:
                                user_content_values.add(str(row_dict['userContent']))
                    
                    return user_content_values
                elif mime in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                              'application/vnd.ms-excel.sheet.macroEnabled.12', 
                              'application/vnd.openxmlformats-officedocument.spreadsheetml.template',
                              'application/vnd.ms-excel.template.macroEnabled.12']:
                    # 如果是 .xlsx 格式但扩展名不对，尝试用 openpyxl 读取
                    from openpyxl import load_workbook
                    wb = None
                    try:
                        wb = load_workbook(filepath, read_only=True)
                        ws = wb.active
                        
                        # 获取所有 userContent 值
                        user_content_values = set()
                        headers = []
                        
                        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                            if row_idx == 0:
                                # 第一行作为列名
                                headers = [str(cell) if cell is not None else f'Column_{i}' for i, cell in enumerate(row)]
                            else:
                                # 数据行
                                row_dict = {headers[i]: (cell if cell is not None else '') for i, cell in enumerate(row)}
                                if 'userContent' in row_dict and row_dict['userContent']:
                                    user_content_values.add(str(row_dict['userContent']))
                        
                        return user_content_values
                    finally:
                        if wb is not None:
                            try:
                                wb.close()
                            except:
                                pass
                else:
                    raise ValueError(f"不支持的文件格式: {file_ext}，MIME类型: {mime}")
            except ImportError:
                # 如果没有安装 python-magic，使用扩展名检测
                raise ValueError(f"不支持的文件格式: {file_ext}。请确保文件是 Excel 格式 (.xlsx, .xls, .xlsm, .xltx, .xltm)")
        except ImportError:
            # 如果 magic 模块不可用，直接使用扩展名检测
            raise ValueError(f"不支持的文件格式: {file_ext}。请确保文件是 Excel 格式 (.xlsx, .xls, .xlsm, .xltx, .xltm)")


@app.route('/api/analyze', methods=['POST'])
def analyze_data():
    """Excel 数据分析接口"""
    try:
        # 清理旧文件
        cleanup_old_files()
        
        # 获取上传的文件
        if 'file' not in request.files:
            return jsonify({'error': '没有上传文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '文件名为空'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': '不支持的文件格式'}), 400
        
        # 保存主文件 - 确保扩展名保留
        filename = secure_filename(file.filename)
        # 确保文件名有扩展名
        if not '.' in filename:
            # 如果 secure_filename 移除了扩展名，手动添加
            original_ext = os.path.splitext(file.filename)[1]
            if original_ext:
                filename += original_ext
            else:
                return jsonify({'error': '文件名没有扩展名'}), 400
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # 检查是否有用于过滤的文件
        filter_file = None
        filter_filepath = None
        if 'filterFile' in request.files:
            filter_file = request.files['filterFile']
            if filter_file.filename != '':
                if not allowed_file(filter_file.filename):
                    return jsonify({'error': '过滤文件格式不支持'}), 400
                filter_filename = secure_filename(filter_file.filename)
                # 确保过滤文件名有扩展名
                if not '.' in filter_filename:
                    # 如果 secure_filename 移除了扩展名，手动添加
                    original_ext = os.path.splitext(filter_file.filename)[1]
                    if original_ext:
                        filter_filename += original_ext
                    else:
                        return jsonify({'error': '过滤文件名没有扩展名'}), 400
                filter_filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"filter_{filter_filename}")
                filter_file.save(filter_filepath)
        
        # 读取主 Excel 文件
        data, headers = read_excel_with_python_libs(filepath)
        
        # 获取参数
        device_ids = request.form.get('deviceIds', '')
        data_types = json.loads(request.form.get('dataTypes', '[]'))
        platform = request.form.get('platform', '安卓')
        
        # 处理设备ID过滤
        device_id_list = [id.strip() for id in device_ids.split(',')] if device_ids else []
        
        # 根据平台过滤
        if platform == '安卓':
            pkg_name = 'com.helloxj.xlook'
        else:
            pkg_name = 'cs.zero.waterCamera'
        
        # 初始数据
        initial_data = [row for row in data if 'pkgName' in row and row['pkgName'] == pkg_name]
        
        # 用户数据（过滤设备ID）
        if device_id_list:
            user_data = [row for row in data 
                        if 'pkgName' in row and row['pkgName'] == pkg_name and 
                        'deviceId' in row and row['deviceId'] not in device_id_list]
        else:
            user_data = [row for row in data if 'pkgName' in row and row['pkgName'] == pkg_name]
        
        # 如果有用于过滤的文件，过滤 userContent 相同的数据
        if filter_filepath:
            filter_user_contents = read_excel_with_python_libs_for_filter(filter_filepath)
            initial_data = [row for row in initial_data 
                           if 'userContent' not in row or str(row['userContent']) not in filter_user_contents]
            user_data = [row for row in user_data 
                        if 'userContent' not in row or str(row['userContent']) not in filter_user_contents]
        
        # 统计数据
        # 初始数据统计
        initial_device_ids = set()
        initial_directives_count = 0
        initial_helpful_count = 0
        initial_unhelpful_count = 0
        
        for row in initial_data:
            if 'deviceId' in row and row['deviceId']:
                initial_device_ids.add(row['deviceId'])
            if 'directives' in row and row['directives'] is not None and str(row['directives']).strip() != '':
                initial_directives_count += 1
            if 'avail' in row and row['avail'] == '有帮助':
                initial_helpful_count += 1
            elif 'avail' in row and row['avail'] == '无帮助':
                initial_unhelpful_count += 1
        
        # 用户数据统计
        user_device_ids = set()
        user_directives_count = 0
        user_helpful_count = 0
        user_unhelpful_count = 0
        
        for row in user_data:
            if 'deviceId' in row and row['deviceId']:
                user_device_ids.add(row['deviceId'])
            if 'directives' in row and row['directives'] is not None and str(row['directives']).strip() != '':
                user_directives_count += 1
            if 'avail' in row and row['avail'] == '有帮助':
                user_helpful_count += 1
            elif 'avail' in row and row['avail'] == '无帮助':
                user_unhelpful_count += 1
        
        initial_stats = {
            '总数据量': len(initial_data),
            '使用人数': len(initial_device_ids),
            '给出指令次数': initial_directives_count,
            '有帮助次数': initial_helpful_count,
            '无帮助次数': initial_unhelpful_count
        }
        
        user_stats = {
            '总数据量': len(user_data),
            '使用人数': len(user_device_ids),
            '给出指令次数': user_directives_count,
            '有帮助次数': user_helpful_count,
            '无帮助次数': user_unhelpful_count
        }
        
        # 构造返回结果
        results = []
        for key in initial_stats.keys():
            results.append({
                'metric': f'{platform} - {key}',
                'initialData': initial_stats[key] if 'initial' in data_types else '',
                'userData': user_stats[key] if 'user' in data_types else ''
            })
        
        # 存储数据到内存数据库供 SQL 查询使用
        data_store['page1_data'] = user_data
        if data_store['conn']:
            data_store['conn'].close()
        data_store['conn'] = sqlite3.connect(':memory:', check_same_thread=False)
        
        # 创建表并插入数据
        cursor = data_store['conn'].cursor()
        
        # 创建表结构（基于数据的列）
        if user_data:
            columns = list(user_data[0].keys())
            # 生成 SQL 创建语句
            create_sql = f"CREATE TABLE data ({', '.join([f'{col} TEXT' for col in columns])})"
            cursor.execute(create_sql)
            
            # 插入数据
            for row in user_data:
                placeholders = ','.join(['?' for _ in columns])
                values = [str(row.get(col, '')) for col in columns]
                cursor.execute(f"INSERT INTO data VALUES ({placeholders})", values)
        
        data_store['conn'].commit()
        
        # 删除上传的文件
        try:
            os.remove(filepath)
        except OSError:
            # 如果文件被占用，跳过删除（后续会被覆盖）
            pass
            
        if filter_filepath:
            try:
                os.remove(filter_filepath)
            except OSError:
                pass

        return jsonify({'data': results})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/label-process', methods=['POST'])
def label_process():
    """小志标签处理接口"""
    try:
        # 清理旧文件
        cleanup_old_files()
        
        # 获取上传的文件
        if 'file' not in request.files:
            return jsonify({'error': '没有上传文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '文件名为空'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': '不支持的文件格式'}), 400
        
        # 保存主文件 - 确保扩展名保留
        filename = secure_filename(file.filename)
        # 确保文件名有扩展名
        if not '.' in filename:
            # 如果 secure_filename 移除了扩展名，手动添加
            original_ext = os.path.splitext(file.filename)[1]
            if original_ext:
                filename += original_ext
            else:
                return jsonify({'error': '文件名没有扩展名'}), 400
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # 检查是否有用于过滤的文件
        filter_file = None
        filter_filepath = None
        if 'filterFile' in request.files:
            filter_file = request.files['filterFile']
            if filter_file.filename != '':
                if not allowed_file(filter_file.filename):
                    return jsonify({'error': '过滤文件格式不支持'}), 400
                filter_filename = secure_filename(filter_file.filename)
                # 确保过滤文件名有扩展名
                if not '.' in filter_filename:
                    # 如果 secure_filename 移除了扩展名，手动添加
                    original_ext = os.path.splitext(filter_file.filename)[1]
                    if original_ext:
                        filter_filename += original_ext
                    else:
                        return jsonify({'error': '过滤文件名没有扩展名'}), 400
                filter_filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"filter_{filter_filename}")
                filter_file.save(filter_filepath)
        
        # 读取 Excel 文件
        data, headers = read_excel_with_python_libs(filepath)
        
        # 检查必需字段
        required_columns = ['question', 'pkgName', 'deviceId']
        if not all(col in headers for col in required_columns):
            return jsonify({'error': f'Excel 文件中缺少必需字段: {", ".join(required_columns)}'}), 400
        
        # 获取参数
        device_ids = request.form.get('deviceIds', '')
        platform = request.form.get('platform', '安卓')
        analysis_type = request.form.get('analysisType', 'default')
        
        # 处理设备ID过滤
        device_id_list = [id.strip() for id in device_ids.split(',')] if device_ids else []
        
        # 根据平台过滤
        if platform == '安卓':
            pkg_name = 'com.helloxj.xlook'
        else:
            pkg_name = 'cs.zero.waterCamera'
        
        # 过滤数据
        if device_id_list:
            filtered_data = [row for row in data 
                            if 'pkgName' in row and row['pkgName'] == pkg_name and 
                            'deviceId' in row and row['deviceId'] not in device_id_list]
        else:
            filtered_data = [row for row in data if 'pkgName' in row and row['pkgName'] == pkg_name]
        
        # 如果有用于过滤的文件，过滤 userContent 相同的数据
        if filter_filepath:
            filter_user_contents = read_excel_with_python_libs_for_filter(filter_filepath)
            filtered_data = [row for row in filtered_data 
                            if 'userContent' not in row or str(row['userContent']) not in filter_user_contents]
        
        # 统计问题/标签出现次数
        question_counts = {}
        for row in filtered_data:
            if 'question' in row and row['question']:
                question = str(row['question'])
                question_counts[question] = question_counts.get(question, 0) + 1
        
        # 根据分析类型过滤
        if analysis_type == 'function':
            # 功能使用：仅展示 question 为功能使用的 userContent 内容
            results = []
            for row in filtered_data:
                if row.get('question', '').strip() == '功能使用':
                    user_content = row.get('userContent', '')
                    if user_content:
                        results.append({'label': user_content, 'count': ''})
        elif analysis_type == 'lowVolume':
            # 低量标签：仅展示标签数小于10，大于等于5的标签名称，但需要展示数量
            filtered_counts = {k: v for k, v in question_counts.items() if 5 <= v < 10}
            results = [{'label': label, 'count': count} for label, count in filtered_counts.items()]
        else:  # 默认分析
            # 默认分析：标签数低于10的统称为其他，数量为低于10的标签数量之和
            other_count = 0
            other_labels = []
            final_counts = {}
            
            for label, count in question_counts.items():
                if count < 10:
                    other_count += count
                    other_labels.append(label)
                else:
                    final_counts[label] = count
            
            # 添加其他项
            if other_labels:
                final_counts['其他'] = other_count
            
            results = [{'label': label, 'count': count} for label, count in final_counts.items()]
            # 按数量排序，其他项放在末尾
            results.sort(key=lambda x: (x['label'] == '其他', -x['count'] if x['label'] != '其他' else 0))
        
        # 存储数据到内存数据库供 SQL 查询使用
        data_store['page2_data'] = filtered_data
        if data_store['conn']:
            data_store['conn'].close()
        data_store['conn'] = sqlite3.connect(':memory:', check_same_thread=False)
        
        # 创建表并插入数据
        cursor = data_store['conn'].cursor()
        
        # 创建表结构（基于数据的列）
        if filtered_data:
            columns = list(filtered_data[0].keys())
            # 生成 SQL 创建语句
            create_sql = f"CREATE TABLE data ({', '.join([f'{col} TEXT' for col in columns])})"
            cursor.execute(create_sql)
            
            # 插入数据
            for row in filtered_data:
                placeholders = ','.join(['?' for _ in columns])
                values = [str(row.get(col, '')) for col in columns]
                cursor.execute(f"INSERT INTO data VALUES ({placeholders})", values)
        
        data_store['conn'].commit()
        
        # 删除上传的文件
        try:
            os.remove(filepath)
        except OSError:
            # 如果文件被占用，跳过删除（后续会被覆盖）
            pass
            
        if filter_filepath:
            try:
                os.remove(filter_filepath)
            except OSError:
                pass
        
        return jsonify({'data': results})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/sql-query', methods=['POST'])
def sql_query():
    """SQL 查询接口"""
    try:
        data = request.get_json()
        query = data.get('query', '')
        
        if not query:
            return jsonify({'error': 'SQL 查询语句不能为空'}), 400
        
        if not data_store['conn']:
            return jsonify({'error': '没有可用的数据，请先加载数据'}), 400
        
        # 执行 SQL 查询
        cursor = data_store['conn'].cursor()
        cursor.execute(query)
        rows = cursor.fetchall()
        
        # 获取列名
        column_names = [description[0] for description in cursor.description]
        
        # 转换为字典列表
        results = [dict(zip(column_names, row)) for row in rows]
        
        return jsonify({'data': results})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/health', methods=['GET'])
def health():
    """健康检查接口"""
    return jsonify({'status': 'ok'})


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
